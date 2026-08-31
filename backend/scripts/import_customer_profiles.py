"""从客服沟通记录表提取客户明细，结构化写入客户画像与需求信号。"""

from __future__ import annotations

import hashlib
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from app.config import get_settings  # noqa: E402
from app.db import SessionLocal  # noqa: E402
from app.models import Customer, DemandSignal  # noqa: E402

SOURCE = "/Users/zhaoxinyuan/Desktop/消费者运营中台/美丽田园/美丽田园cod/美丽田园项目资料/美丽田园客服日常沟通记录表.xlsx"


def _clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"-", "否", "/", "未知", "None", "nan"}:
        return ""
    return text


def _split_tags(value) -> list[str]:
    tags: list[str] = []
    for item in re.split(r"[,，;；]", _clean(value)):
        item = item.strip()
        if item and item not in {"否", "-"}:
            tags.append(item)
    return tags


def _phone_hash(phone: str) -> str:
    digits = re.sub(r"\D", "", phone)
    return hashlib.sha1(digits.encode("utf-8")).hexdigest()[:16] if digits else ""


def _one_id(phone: str, external_id: str, name: str) -> str:
    key = phone or external_id or name
    if not key:
        return ""
    return "mly-" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


def _headers(rows, header_row_idx: int) -> dict[str, int]:
    header = rows[header_row_idx]
    mapping: dict[str, int] = {}
    for idx, value in enumerate(header):
        name = _clean(value)
        if name and name not in mapping:
            mapping[name] = idx
    return mapping


def _row_dict(row, mapping: dict[str, int]) -> dict[str, str]:
    return {name: _clean(row[idx] if idx < len(row) else "") for name, idx in mapping.items()}


def _merge_tags(profile: dict, *tag_sources: list[str]) -> list[str]:
    existing = set(profile.get("tags", []))
    for source in tag_sources:
        existing.update(_split_tags(source))
    return sorted(existing)


def main() -> None:
    settings = get_settings()
    tenant_id = settings.default_tenant_id
    customers: dict[str, Customer] = {}
    signals: list[tuple[Customer, str]] = []

    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    try:
        # 20元优惠券客户信息触达表：最全的客户标签
        ws = wb["20元优惠券客户信息触达表"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        mapping = {name: idx for idx, name in enumerate(header) if name}
        for row in rows:
            data = {name: _clean(row[idx] if idx < len(row) else "") for name, idx in mapping.items()}
            phone = data.get("电话", "")
            name = data.get("客户名称") or data.get("备注名")
            external_id = data.get("external_userid") or data.get("unionid")
            oid = _one_id(phone, external_id, name)
            if not oid or not name:
                continue
            profile: dict = {"tags": [], "purchases": [], "demands": []}
            profile["tags"] = _merge_tags(
                profile,
                data.get("个人标签", ""),
                data.get("企业标签", ""),
                data.get("标签组(皮肤需求)", ""),
                data.get("标签组(皮肤类型)", ""),
                data.get("标签组(vip等级)", ""),
                data.get("标签组(客户卡等级)", ""),
                data.get("标签组(客户年龄)", ""),
                data.get("标签组(区域明细)", ""),
                data.get("标签组(健康状态)", ""),
                data.get("标签组(活动内容标签)", ""),
                data.get("标签组(客服SOP进度)", ""),
            )
            profile["phone_hash"] = _phone_hash(phone)
            profile["city"] = data.get("城市/地区", "")
            profile["age"] = data.get("年龄", "")
            profile["gender"] = data.get("性别", "")
            profile["status"] = data.get("客户状态") or data.get("流失状态", "")
            profile["channel"] = data.get("添加渠道") or data.get("来源", "")
            profile["added_at"] = data.get("添加时间", "")
            profile["last_communication"] = data.get("最近沟通时间", "")
            profile["vip_level"] = data.get("标签组(vip等级)", "") or data.get("标签组(客户卡等级)", "")
            profile["skin_needs"] = _split_tags(data.get("标签组(皮肤需求)", ""))
            customer = Customer(
                tenant_id=tenant_id,
                one_id=oid,
                name=name[:80],
                profile_json=profile,
            )
            customers[oid] = customer

        # 3月在职继承：补充标签
        ws = wb["3月在职继承"]
        rows = list(ws.iter_rows(values_only=True))
        mapping = _headers(rows, header_row_idx=4)
        for row in rows[5:]:
            data = _row_dict(row, mapping)
            name = data.get("客户名称") or data.get("备注名")
            oid = _one_id(data.get("电话", ""), "", name)
            if not oid or not name:
                continue
            customer = customers.setdefault(
                oid,
                Customer(
                    tenant_id=tenant_id,
                    one_id=oid,
                    name=name[:80],
                    profile_json={"tags": [], "purchases": [], "demands": []},
                ),
            )
            profile = dict(customer.profile_json or {})
            profile["tags"] = _merge_tags(
                profile,
                data.get("企业标签", ""),
                data.get("标签组(客户卡等级)", ""),
                data.get("标签组(私域引流方式)", ""),
                data.get("标签组(私域流量来源)", ""),
            )
            profile["channel"] = data.get("添加时间", "") and profile.get("channel", "") or profile.get("channel", "")
            profile["added_at"] = data.get("添加时间", "") or profile.get("added_at", "")
            profile["is_ordered"] = data.get("是否下单", "")
            customer.profile_json = profile

        # 卖卡销售：购买记录
        ws = wb["卖卡销售"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        mapping = {name: idx for idx, name in enumerate(header) if name}
        purchase_count = 0
        for row in rows:
            data = {name: _clean(row[idx] if idx < len(row) else "") for name, idx in mapping.items()}
            phone = data.get("手机号", "")
            nickname = data.get("微信昵称", "")
            oid = _one_id(phone, "", nickname)
            customer = customers.get(oid)
            if customer is None:
                customer = Customer(
                    tenant_id=tenant_id,
                    one_id=oid,
                    name=nickname[:80],
                    profile_json={"tags": [], "purchases": [], "demands": []},
                )
                customers[oid] = customer
            profile = dict(customer.profile_json or {})
            purchases = list(profile.get("purchases", []))
            card = data.get("购买卡项", "")
            amount = data.get("销售金额", "")
            if card or amount:
                purchases.append(
                    {
                        "date": data.get("下单时间", ""),
                        "card": card,
                        "amount": amount,
                        "source": data.get("客户来源", ""),
                        "appointment": data.get("预约时间", ""),
                        "channel": data.get("来源渠道", ""),
                        "count": data.get("成单数", ""),
                    }
                )
                purchase_count += 1
            profile["purchases"] = purchases[-50:]
            customer.profile_json = profile

        # 回复未购意向用户：需求信号
        ws = wb["回复未购意向用户"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        mapping = {name: idx for idx, name in enumerate(header) if name}
        demand_count = 0
        for row in rows:
            data = {name: _clean(row[idx] if idx < len(row) else "") for name, idx in mapping.items()}
            phone = data.get("电话", "")
            nickname = data.get("用户昵称", "")
            demand_text = data.get("用户需求", "")
            oid = _one_id(phone, "", nickname)
            customer = customers.get(oid)
            if customer is None:
                customer = Customer(
                    tenant_id=tenant_id,
                    one_id=oid,
                    name=nickname[:80],
                    profile_json={"tags": [], "purchases": [], "demands": []},
                )
                customers[oid] = customer
            profile = dict(customer.profile_json or {})
            demands = list(profile.get("demands", []))
            demands.append(
                {
                    "date": data.get("沟通时间", ""),
                    "demand": demand_text[:500],
                    "source": data.get("来源渠道", ""),
                    "card": data.get("沟通卡项", ""),
                    "user_attribute": data.get("用户属性（体验客/新客）", ""),
                }
            )
            profile["demands"] = demands[-50:]
            profile["tags"] = _merge_tags(profile, data.get("用户属性（体验客/新客）", ""))
            customer.profile_json = profile
            if demand_text:
                signals.append((customer, demand_text[:1000]))
                demand_count += 1
    finally:
        wb.close()

    with SessionLocal() as db:
        existing_ids = {
            c.one_id for c in db.query(Customer).filter(Customer.tenant_id == tenant_id).all()
        }
        new_customers = [c for c in customers.values() if c.one_id not in existing_ids]
        db.add_all(new_customers)
        db.flush()
        db.add_all(
            DemandSignal(
                tenant_id=tenant_id,
                customer_id=customer.id,
                source_type="cs",
                raw_content=text,
            )
            for customer, text in signals
        )
        db.commit()

        total = db.query(Customer).filter(Customer.tenant_id == tenant_id).count()
        total_signals = db.query(DemandSignal).filter(DemandSignal.tenant_id == tenant_id).count()
        print("===== 客户画像导入汇总 =====")
        print(f"本次新增客户: {len(new_customers)}")
        print(f"累计客户总数: {total}")
        print(f"购买记录: {purchase_count}")
        print(f"需求信号: {demand_count}，累计信号: {total_signals}")


if __name__ == "__main__":
    main()
