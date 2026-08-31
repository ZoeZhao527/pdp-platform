"""超大品牌资料分批导入：xlsx 按 sheet+批次，PDF 按整册。"""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from pathlib import Path
from typing import Iterator
from xml.etree import ElementTree as ET

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.config import get_settings  # noqa: E402
from app.db import SessionLocal  # noqa: E402
from app.knowledge.service import KnowledgeService  # noqa: E402
from app.models import KnowledgeDoc  # noqa: E402

ROOT = Path("/Users/zhaoxinyuan/Desktop/消费者运营中台/美丽田园")
MAX_ROWS_PER_PART = 2000
MAX_COLS = 80

LARGE_FILES = [
    "美丽田园cod/美丽田园项目资料/sms总表.xlsx",
    "美丽田园cod/美丽田园项目资料/内容向.xlsx",
    "美丽田园cod/美丽田园项目资料/美丽田园项目进度表.xlsx",
    "美丽田园cod/美丽田园项目资料/美丽田园客服日常沟通记录表.xlsx",
    "美丽田园cod/美丽田园项目资料/美丽田园基础资料 2/美丽田园基础资料/2025美丽田园疗程手册1010.pdf",
    "美丽田园cod/美丽田园项目资料/美丽田园基础资料 2/美丽田园基础资料/2025美丽田园疗程手册1118.pdf",
    "美丽田园cod/美丽田园项目资料/美丽田园基础资料 2/美丽田园基础资料/美丽田园集团介绍25.pdf",
    "美丽田园cod/美丽田园项目资料/美丽田园基础资料 2/美丽田园基础资料/美丽田园VI视觉识别系统（2023升级）.pdf",
]


def _redact(text: str) -> str:
    text = re.sub(r"1[3-9]\d{9}", "[手机号已脱敏]", text)
    text = re.sub(r"wxid_[A-Za-z0-9_-]+", "[微信号已脱敏]", text)
    text = re.sub(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", "[邮箱已脱敏]", text)
    return text


def _line(values: list) -> str:
    cleaned = [str(v).strip() for v in values[:MAX_COLS] if v is not None and str(v).strip()]
    return " | ".join(cleaned)


def _parts_from_lines(sheet: str, lines: list[str]) -> Iterator[tuple[str, int, str]]:
    part = 1
    for start in range(0, len(lines), MAX_ROWS_PER_PART):
        chunk = lines[start : start + MAX_ROWS_PER_PART]
        if chunk:
            yield sheet, part, "\n".join(chunk)
        part += 1


def iter_xlsx_parts(path: Path) -> Iterator[tuple[str, int, str]]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            lines = [line for line in (_line(row) for row in ws.iter_rows(values_only=True)) if line]
            yield from _parts_from_lines(ws.title, lines)
        wb.close()
        return
    except Exception:  # noqa: BLE001
        pass

    # openpyxl 无法解析的样式损坏文件，走底层 XML 解析
    NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path) as z:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in root.findall("m:si", NS):
                shared.append("".join(t.text or "" for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")))
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        sheets = [s.get("name") for s in wb.find("m:sheets", NS)]
        for idx, sheet in enumerate(sheets, start=1):
            xml_path = f"xl/worksheets/sheet{idx}.xml"
            if xml_path not in z.namelist():
                continue
            root = ET.fromstring(z.read(xml_path))
            lines: list[str] = []
            for row in root.findall(".//m:sheetData/m:row", NS):
                values: list[str] = []
                for cell in row.findall("m:c", NS):
                    cell_type = cell.get("t")
                    value_el = cell.find("m:v", NS)
                    if cell_type == "s" and value_el is not None:
                        values.append(shared[int(value_el.text)])
                    elif cell_type == "inlineStr":
                        inline = cell.find("m:is", NS)
                        values.append(
                            "".join(t.text or "" for t in inline.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"))
                            if inline is not None
                            else ""
                        )
                    elif value_el is not None:
                        values.append(value_el.text or "")
                line = _line(values)
                if line:
                    lines.append(line)
            yield from _parts_from_lines(sheet, lines)


def extract_pdf(path: Path) -> str:
    import pdfplumber

    parts: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text.strip():
                parts.append(text)
    return "\n\n".join(parts)


def main() -> None:
    settings = get_settings()
    with SessionLocal() as db:
        service = KnowledgeService(db)
        existing = {
            row.name
            for row in db.query(KnowledgeDoc).filter(KnowledgeDoc.tenant_id == settings.default_tenant_id).all()
        }
        seen_hashes: set[str] = set()
        imported = 0
        failed: list[str] = []

        for rel in LARGE_FILES:
            path = ROOT / rel
            if not path.exists():
                print(f"[缺失] {rel}")
                continue
            stem = re.sub(r"[\s]+", "", path.stem)
            try:
                if path.suffix.lower() == ".xlsx":
                    chunks: list[tuple[str, int, str]] = list(iter_xlsx_parts(path))
                    for sheet, part, text in chunks:
                        text = _redact(text).strip()
                        if not text:
                            continue
                        digest = hashlib.md5(text.encode("utf-8")).hexdigest()
                        if digest in seen_hashes:
                            continue
                        seen_hashes.add(digest)
                        name = f"美丽田园-大文件-{stem}-{sheet}-part{part}.xlsx"[:200]
                        if name in existing:
                            continue
                        doc = service.ingest(settings.default_tenant_id, name, text, "xlsx", path.stat().st_size)
                        existing.add(name)
                        imported += 1
                        print(f"[导入] {name} -> {doc.chunk_count} 切片")
                else:
                    text = _redact(extract_pdf(path)).strip()
                    digest = hashlib.md5(text.encode("utf-8")).hexdigest()
                    if digest in seen_hashes:
                        print(f"[跳过-重复] {rel}")
                        continue
                    seen_hashes.add(digest)
                    name = f"美丽田园-大文件-{stem}.pdf"[:200]
                    if name in existing:
                        continue
                    doc = service.ingest(settings.default_tenant_id, name, text, "pdf", path.stat().st_size)
                    existing.add(name)
                    imported += 1
                    print(f"[导入] {name} -> {doc.chunk_count} 切片（文字 {len(text)} 字）")
            except Exception as exc:  # noqa: BLE001
                failed.append(f"{rel}: {exc}")
                print(f"[失败] {rel}: {exc}")

        print(f"\n===== 大文件导入汇总 =====")
        print(f"导入成功: {imported}")
        print(f"失败: {len(failed)}")
        for item in failed:
            print(f"  - {item}")


if __name__ == "__main__":
    main()

